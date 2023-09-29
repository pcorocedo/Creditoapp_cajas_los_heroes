import time
import pandas as pd
import pyodbc
import logging
import os
import shutil
import openpyxl
import datetime
from shutil import rmtree
from tqdm.auto import tqdm

logging.basicConfig(filename="data.log", level=logging.DEBUG,
                    format='%(asctime)s :: %(levelname)s :: %(funcName)s :: %(lineno)d :: %(message)s')
logging.debug("_____________________tablas______________________")
print("_____________________tablas______________________")

logging.info("creacion de tablas A.4")
logging.info("creacion de tablas A.4")
A_4_Registro_Detalle = [{
    'Tipo_Registro': ' ',
    'Rut_del_trabajador': ' ',
    'Digito_Verificador_del_Trabajador': ' ',
    'Apellido_paterno_del_Trabajador': ' ',
    'Apellido_materno_trabajador': ' ',
    'Nombres_del_trabajador': ' ',
    'Sexo': ' ',
    'Institución_de_Salud': ' ',
    'Renta_Imponible_del_trabajador': ' ',
    'Días_Trabajados': ' ',
    'Cantidad_de_Cargas_Simples': ' ',
    'Cantidad_de_Cargas_Invalidas': ' ',
    'Cantidad_de_Cargas_Maternales': ' ',
    'Monto_de_Asignación_Familiar': ' ',
    'Monto_de_Creditos_Personales': ' ',
    'Monto_de_Convenios_Dentales': ' ',
    'Monto_Leasing': ' ',
    'Montos_de_Seguros_de_Vidas': ' ',
    'Aporte_1': ' ',
    'Aporte_Adicional': ' ',
    'Otros_CCAF': ' ',
    'Tramo_de_Asignacion_Familiar': ' ',
    'Monto_Asignación_Familiar_Retroactiva': ' ',
    'Monto_Reintegros_Asignación_Familiar': ' ',
    'Codigo_de_Movimiento_de_Personal': ' ',
    'Fecha_Inicio_de_Movimiento_de_Personal': ' ',
    'Fecha_Término_de_Movimiento_de_Personal': ' ',
    'Periodo_Pago': ' ',
    'Monto_6_de_trabajador_no_afiliado_a_Isapre': ' ',
    'Renta_Imponible_diaria_del_trabajador': ' ',
    'Total_días_con_subsidio': ' ',
    'Código_de_licencia_Medica': ' ',
    'Fecha_con_Subsidio_desde': ' ',
    'Fecha_con_Subsidio_hasta': ' ',
    'Nacionalidad': ' ',
    'Filler': ' '
}]
servipag = [{
    'Tipo_Registro': ' ',
    'Rut_del_trabajador': ' ',
    'Digito_Verificador_del_Trabajador': ' ',
    'Apellido_paterno_del_Trabajador': ' ',
    'Apellido_materno_trabajador': ' ',
    'Nombres_del_trabajador': ' ',
    'Sexo': ' ',
    'Institución_de_Salud': ' ',
    'Renta_Imponible_del_trabajador': ' ',
    'Días_Trabajados': ' ',
    'Cantidad_de_Cargas_Simples': ' ',
    'Cantidad_de_Cargas_Invalidas': ' ',
    'Cantidad_de_Cargas_Maternales': ' ',
    'Monto_de_Asignación_Familiar': ' ',
    'Monto_de_Creditos_Personales': ' ',
    'Monto_de_Convenios_Dentales': ' ',
    'Monto_Leasing': ' ',
    'Montos_de_Seguros_de_Vidas': ' ',
    'Aporte_1': ' ',
    'Aporte_Adicional': ' ',
    'Otros_CCAF': ' ',
    'Tramo_de_Asignacion_Familiar': ' ',
    'Monto_Asignación_Familiar_Retroactiva': ' ',
    'Monto_Reintegros_Asignación_Familiar': ' ',
    'Codigo_de_Movimiento_de_Personal': ' ',
    'Fecha_Inicio_de_Movimiento_de_Personal': ' ',
    'Fecha_Término_de_Movimiento_de_Personal': ' ',
    'Periodo_Pago': ' ',
    'Monto_6_de_trabajador_no_afiliado_a_Isapre': ' ',
    'Renta_Imponible_diaria_del_trabajador': ' ',
    'Total_días_con_subsidio': ' ',
    'Código_de_licencia_Medica': ' ',
    'Fecha_con_Subsidio_desde': ' ',
    'Fecha_con_Subsidio_hasta': ' ',
    'Nacionalidad': ' ',
    'Filler': ' '
}]
index = [{
    'Tipo_Registro': ' ',
    'Rut_del_trabajador': ' ',
    'Digito_Verificador_del_Trabajador': ' ',
    'Apellido_paterno_del_Trabajador': ' ',
    'Apellido_materno_trabajador': ' ',
    'Nombres_del_trabajador': ' ',
    'Sexo': ' ',
    'Institución_de_Salud': ' ',
    'Renta_Imponible_del_trabajador': ' ',
    'Días_Trabajados': ' ',
    'Cantidad_de_Cargas_Simples': ' ',
    'Cantidad_de_Cargas_Invalidas': ' ',
    'Cantidad_de_Cargas_Maternales': ' ',
    'Monto_de_Asignación_Familiar': ' ',
    'Monto_de_Creditos_Personales': ' ',
    'Monto_de_Convenios_Dentales': ' ',
    'Monto_Leasing': ' ',
    'Montos_de_Seguros_de_Vidas': ' ',
    'Aporte_1': ' ',
    'Aporte_Adicional': ' ',
    'Otros_CCAF': ' ',
    'Tramo_de_Asignacion_Familiar': ' ',
    'Monto_Asignación_Familiar_Retroactiva': ' ',
    'Monto_Reintegros_Asignación_Familiar': ' ',
    'Codigo_de_Movimiento_de_Personal': ' ',
    'Fecha_Inicio_de_Movimiento_de_Personal': ' ',
    'Fecha_Término_de_Movimiento_de_Personal': ' ',
    'Periodo_Pago': ' ',
    'Monto_6_de_trabajador_no_afiliado_a_Isapre': ' ',
    'Renta_Imponible_diaria_del_trabajador': ' ',
    'Total_días_con_subsidio': ' ',
    'Código_de_licencia_Medica': ' ',
    'Fecha_con_Subsidio_desde': ' ',
    'Fecha_con_Subsidio_hasta': ' ',
    'Nacionalidad': ' ',
    'Filler': ' '
}]
logging.debug("_____________________variable locales_____________")
print("_____________________variable locales_____________")
def config (variable):
    """
    #Consulta de variable del libro de excel "Data/Config.xlsx"
    #ejemplo solo config(el nombre de la variable)
    """
    configuracion = pd.read_excel("Data/Config.xlsx",sheet_name='Variables ')
    (columna , celdas) = configuracion.shape
    x =-1
    while x < int(columna):
        x = 1+x
        if configuracion.iloc[int(x),0] == variable : break
    return configuracion.iloc[int(x),(1)]

logging.debug("_____________________        Querys   _____________")
def QUERY_NOMINA_COBRO_TRAB_VCTO(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select * from NOMINA_COBRO_TRAB_VCTO_20230510_carga where rut_pens_tran = 9094715

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    QUERY_NOMINA_COBRO_TRAB_VCTO(fecha,rut)

    """
    query = str(config("QUERY_NOMINA_COBRO_TRAB_VCTO")).format(fecha, rut)
    return query
def NOMINA_SEGURO(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select rut_pens_tran,Rut_Empresa,Dig_Ver_Compania,Monto_Desc_Valor  from NOMINA_SEGURO_082023pensionado
where rut_pens_tran = 6904044 AND enviado = 'si'

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    NOMINA_SEGURO_(fecha,rut)

    """
    query = str(config("NOMINA_SEGURO")).format(fecha, rut)
    return query
def nomina_hipo_trab_vcto(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select rut_pens_trab,Rut_Empresa,Monto_Desc_Valor from nomina_hipo_trab_vcto_20230310

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    nomina_hipo_trab_vcto(fecha,rut)

    """
    query = str(config("nomina_hipo_trab_vcto")).format(fecha, rut)
    return query
def NOMINA_COBRO_pens_VCTO_(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select * from NOMINA_COBRO_TRAB_VCTO_20230510_carga where rut_pens_tran = 9094715

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    QUERY_NOMINA_COBRO_TRAB_VCTO(fecha,rut)

    """
    query = str(config("NOMINA_COBRO_pens_VCTO_")).format(fecha, rut)
    return query
def NOMINA_SEGURO_PENSIONADO(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select rut_pens_tran,Rut_Empresa,Dig_Ver_Compania,Monto_Desc_Valor  from NOMINA_SEGURO_082023pensionado
where rut_pens_tran = 6904044 AND enviado = 'si'

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    NOMINA_SEGURO_(fecha,rut)

    """
    query = str(config("NOMINA_SEGURO_PENSIONADO")).format(fecha, rut)
    return query
def nomina_hipo_pens_vcto_(fecha, rut):
    """
    Esta funcion representa la consulta por ejemplo:
select rut_pens_trab,Rut_Empresa,Monto_Desc_Valor from nomina_hipo_trab_vcto_20230310

    :param fecha:
    :param rut:
    :return:
    _______________________________________________________________________________
    ejemplo:
    nomina_hipo_trab_vcto(fecha,rut)

    """
    query = str(config("nomina_hipo_pens_vcto_")).format(fecha, rut)
    return query

logging.debug("_____________________Funciones____________________")
print("_____________________Tareas   ____________________")
def conexionBD():
    """
    datos para conectar la base datos con pyodbc los datos ejemplo
        server : 'tcp:myserver.database.windows.net'
        database : 'mydb'
        username : 'myusername'
        password : 'mypassword'
    """
    try:
        server = str(config ("server"))
        database = str(config ("database"))
        username = str(config ("username"))
        password = str(config ("pass"))


        connection =pyodbc.connect(
            'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server +
            ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)


        logging.info("Conectando ala base datos con exito")
        return connection
    except:
         logging.error("Error en conectar , la base de datos ")
         print("Error en conectar , la base de datos ")
def PL():
  """
        modelo que extrae todos los datos de archivo txt que inicia con "PL":
        Archivo : archivo a consultar
        la salida del modelo son :
            * PagosElectronicos app\SalidaCVS
            *PagosElectronicos app\SalidaExcel

  :return: lectura de todos los PL
  """
  try:
      path="Data/inputPL.txt"
      logging.info("Abriendo Txt que inicia con PL")
      f = open(path, "r")
      for x in f:
        Tipo_Registro = str(x[1:1])
        Rut_del_trabajador = str(x[2:10])
        Digito_Verificador_del_Trabajador = str(x[10:11])
        Apellido_paterno_del_Trabajador = str(x[11:41])
        Apellido_materno_trabajador = str(x[41:71])
        Nombres_del_trabajador = str(x[71:101])
        Sexo = str(x[101:102])
        Institución_de_Salud = str(x[102:104])
        Renta_Imponible_del_trabajador = str(x[104:116])
        Días_Trabajados = str(x[116:128])
        Cantidad_de_Cargas_Simples = str(x[128:140])
        Cantidad_de_Cargas_Invalidas = str(x[140:152])
        Cantidad_de_Cargas_Maternales = str(x[152:164])
        Monto_de_Asignación_Familiar = str(x[164:176])
        Monto_de_Creditos_Personales = str(x[176:188])
        Monto_de_Convenios_Dentales = str(x[188:200])
        Monto_Leasing = str(x[200:212])
        Montos_de_Seguros_de_Vidas = str(x[212:224])
        Aporte_1 = str(x[224:236])
        Aporte_Adicional = str(x[236:248])
        Otros_CCAF = str(x[248:260])
        Tramo_de_Asignacion_Familiar = str(x[260:261])
        Monto_Asignación_Familiar_Retroactiva = str(x[261:273])
        Monto_Reintegros_Asignación_Familiar = str(x[273:285])
        Codigo_de_Movimiento_de_Personal = str(x[285:289])
        Fecha_Inicio_de_Movimiento_de_Personal = str(x[289:297])
        Fecha_Término_de_Movimiento_de_Personal = str(x[297:305])
        Periodo_Pago = str(x[305:311])
        Monto_6_de_trabajador_no_afiliado_a_Isapre = str(x[311:323])
        Renta_Imponible_diaria_del_trabajador = str(x[323:335])
        Total_días_con_subsidio = str(x[335:339])
        Código_de_licencia_Medica = str(x[339:340])
        Fecha_con_Subsidio_desde = str(x[340:348])
        Fecha_con_Subsidio_hasta = str(x[348:356])
        Nacionalidad = str(x[356:357])
        Filler = str(x[357:754])

    #calculadon  los productos financieros
        try:
            afa = str(int(Monto_6_de_trabajador_no_afiliado_a_Isapre) - int(Monto_Asignación_Familiar_Retroactiva) - int(Monto_Reintegros_Asignación_Familiar)-int(Monto_de_Asignación_Familiar))
            Productos_financieros = str(int(Monto_de_Creditos_Personales) + int(Monto_de_Convenios_Dentales) +int(Monto_Leasing) +int(Montos_de_Seguros_de_Vidas) +int(Otros_CCAF))
        except:
            Productos_financieros = None
            afa=None
            pass

        if Productos_financieros != None and  Rut_del_trabajador != "00000000" :
            if int(afa) < 0:
                afa=0
                A_4_Registro_Detalle.append({
                    'Tipo_Registro': Tipo_Registro,
                    'Rut_del_trabajador': int(Rut_del_trabajador),
                    'Digito_Verificador_del_Trabajador': Digito_Verificador_del_Trabajador,
                    'rut':Rut_del_trabajador+ " - " +Digito_Verificador_del_Trabajador,
                    'Apellido_paterno_del_Trabajador': Apellido_paterno_del_Trabajador,
                    ' Apellido_materno_trabajador': Apellido_materno_trabajador,
                    'Nombres_del_trabajador': Nombres_del_trabajador,
                    'Sexo': Sexo,
                    'Institución_de_Salud': Institución_de_Salud,
                    'Renta_Imponible_del_trabajador': Renta_Imponible_del_trabajador,
                    'Días_Trabajados': Días_Trabajados,
                    'Cantidad_de_Cargas_Simples': Cantidad_de_Cargas_Simples,
                    'Cantidad_de_Cargas_Invalidas': Cantidad_de_Cargas_Invalidas,
                    'Cantidad_de_Cargas_Maternales': Cantidad_de_Cargas_Maternales,
                    'Monto_de_Asignación_Familiar': int(Monto_de_Asignación_Familiar),
                    'Monto_de_Creditos_Personales': int(Monto_de_Creditos_Personales),
                    'Monto_de_Convenios_Dentales': int(Monto_de_Convenios_Dentales),
                    'Monto_Leasing': int(Monto_Leasing),
                    'Montos_de_Seguros_de_Vidas': int(Montos_de_Seguros_de_Vidas),
                    'Aporte_1': int(Aporte_1),
                    'Aporte_Adicional': int(Aporte_Adicional),
                    'Otros_CCAF': int(Otros_CCAF),
                    'Tramo_de_Asignacion_Familiar': int(Monto_de_Asignación_Familiar),
                    'Monto_Asignación_Familiar_Retroactiva': int(Monto_Asignación_Familiar_Retroactiva),
                    'Monto_Reintegros_Asignación_Familiar': int(Monto_Reintegros_Asignación_Familiar),
                    'Codigo_de_Movimiento_de_Personal': Codigo_de_Movimiento_de_Personal,
                    'Fecha_Inicio_de_Movimiento_de_Personal': Fecha_Inicio_de_Movimiento_de_Personal,
                    'Fecha_Término_de_Movimiento_de_Personal': Fecha_Término_de_Movimiento_de_Personal,
                    'Periodo_Pago': Periodo_Pago,
                    'Monto_6_de_trabajador_no_afiliado_a_Isapre': int(Monto_6_de_trabajador_no_afiliado_a_Isapre),
                    'Renta_Imponible_diaria_del_trabajador': Renta_Imponible_diaria_del_trabajador,
                    'Total_días_con_subsidio': Total_días_con_subsidio,
                    'Código_de_licencia_Medica': Código_de_licencia_Medica,
                    'Fecha_con_Subsidio_desde': Fecha_con_Subsidio_desde,
                    'Fecha_con_Subsidio_hasta': Fecha_con_Subsidio_hasta,
                    'Nacionalidad': Nacionalidad,
                    'Filler': Filler,
                    'Productos_financieros': int(Productos_financieros),
                    'afa': int(afa),


                    })

      logging.info("Imprimiendo resultados")
      resultado = pd.DataFrame(A_4_Registro_Detalle)

      resultado = resultado[['Rut_del_trabajador',
                               'rut',
                               'Monto_de_Asignación_Familiar',
                               'Monto_de_Creditos_Personales',
                               'Monto_de_Convenios_Dentales',
                               'Monto_Leasing',
                               'Montos_de_Seguros_de_Vidas',
                               'Aporte_1', 'Aporte_Adicional',
                               'Otros_CCAF',
                               'Monto_Asignación_Familiar_Retroactiva',
                               'Monto_Reintegros_Asignación_Familiar',
                               'Monto_6_de_trabajador_no_afiliado_a_Isapre',
                               'Productos_financieros',
                               'afa'
                               ]]

      resultado.to_csv('SalidaCVS//index.csv', encoding='utf-8')
      resultado.to_excel('SalidaExcel//A_4_Registro_Detalle.xlsx', sheet_name='A4', index=False, header=True)
      print("lectura de PL's terminada...")
      for i in tqdm(range(20)):
          # Apoyo visual pasa saber el progreso del bot ser borrado  y no afectara el codigo
          time.sleep(000000000000000.1)
          print(" ", end='\r')
      return resultado
  except:
      print("Error en lectura de PL")
      logging.error("Error en lectura de PL")
def insertar_datos_sql():
    """
    funcion que inserta valores en SQLserve:

    """
    connection = conexionBD()
    #stringquery = "INSERT INTO A4_TABLA VALUES ({}, '{}', {},{},{},{},{},{},{},{},{},{},{});".format(2222222, 2, 3000,500,500,500,500,500,500,500,500,500,500)
    strinqquery = "INSERT INTO A4_TABLA VALUES (33333333, '2', 3000,500,500,500,500,500,500,500,500,500,500);"
    cursor = connection.cursor()
    cursor.execute(strinqquery)
    cursor.commit()
    cursor.close()
def Select_datos_sql(query):
    """
    funcion para leer tabla  valores en SQLserve:
    debe consultar el config.xlsx que esta en el archivo de excel en la carpeta
    data por ejemplo :
    ____________________________________________________________________________
    Select_datos_sql(NOMINA_COBRO_TRAB_VCTO )
    ____________________________________________________________________________
    """
    try:
        connection = conexionBD()
        cursor = connection.cursor()
        Dt= cursor.execute(query)
        return Dt
        cursor.close()
    except:
        logging.error("Error en el select revisar la funcion def Select_datos_sql(query) ")
def eliminarcarpetas():
    try:
        rmtree("SalidaCVS")
        rmtree("SalidaExcel")
        rmtree("SalidaSQL")
        os.remove("Data/inputPL.txt")
        os.remove("Data/inputHE.txt")

        logging.info("Eliminamos carpetas y archivos ")
        print("Eliminamos carpetas y archivos ")

    except:
        logging.error("error en la Eliminacion de carpetas")
        print("error en la Eliminacion de carpetas")
        pass
    finally:

        for i in tqdm(range(20)):
            # Apoyo visual pasa saber el progreso del bot , pude ser borrado  y no afectara el codigo
            time.sleep(000000000000000.1)
            print(" ", end='\r')
def creacionCarpetas():
    logging.info("Creado las carpetas ")
    print("Creado las carpetas ")

    try:
        logging.info("Creado las carpetas ")
        os.mkdir('SalidaCVS')
        os.mkdir('SalidaExcel')
        os.mkdir("SalidaSQL")
        w = open("Data/inputPL.txt", "a")
        w.close()
    except:
        logging.error("error en la creacion de carpetas")
        print("error en la creacion de carpetas")
        pass
    finally:

        for i in tqdm(range(20)):
            # Apoyo visual pasa saber el progreso del bot , pude ser borrado  y no afectara el codigo
            time.sleep(000000000000000.1)
            print(" ", end='\r')
def ETLA4_TRABAJADOR():
    """
    ETL que hace las consultas en:
    1. Consulta Facturado Trabajador Crédito.
    2. Consulta Facturado Trabajador Seguros
    3. Consulta Facturado Trabajador Hipotecario
    :return: Prelacion previred.
    """

    listaETL = [{}]
    def Facturado_Crédito():
        print("Consulta Facturado Trabajador Crédito")
        ETL=pd.read_csv("SalidaCVS/index.csv")

        # Consulta Facturado Trabajador Crédito
        for i in tqdm(range(len(ETL))):

                try:
                    dtable = Select_datos_sql(
                        QUERY_NOMINA_COBRO_TRAB_VCTO(str(config("Consulta")),
                                                     ETL.iloc[i]['Rut_del_trabajador'])
                    )
                    df = pd.DataFrame(dtable)
                    rut_pens_tran, Rut_Empresa, Dig_Ver_Compania, Monto_Desc_Valor = df.iloc[0, 0]

                except:
                    rut_pens_tran = 0
                    Rut_Empresa = 0
                    Dig_Ver_Compania = 0
                    Monto_Desc_Valor = 0
                    if rut_pens_tran == 0 and Rut_Empresa == 0 and Dig_Ver_Compania == 0 and Monto_Desc_Valor == 0:
                        rut_pens_tran = ""
                        Rut_Empresa = ""
                        Dig_Ver_Compania = ""
                        Monto_Desc_Valor =0
                    pass
                finally:

                    listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['Rut_del_trabajador'],
                              'RUT_EMPRESA': Rut_Empresa,
                               'DV_EMPRESA': Dig_Ver_Compania,
                               'MONTO': ETL.iloc[i]['Productos_financieros'],
                               'CRUCE CON NOMINA DE CREDITO': Monto_Desc_Valor,
                               'CRUCE CON NOMINA DE SEGUROS': ' ',
                               'CRUCE HIPOTECARIO': ' ',
                               'DIFERENCIAS': ' ',
                              })

                print(" ", end='\r')#Consulta Facturado Trabajador Crédito #Consulta Facturado Trabajador Crédito
        lETLpaso1 = pd.DataFrame(listaETL)
        lETLpaso1.to_csv("SalidaCVS/ETL01.csv")
    def Facturado_Seguros():
        # Consulta Facturado Trabajador Seguros
        print("Consulta Facturado Trabajador Seguros")
        ETL = pd.read_csv("SalidaCVS/ETL01.csv", encoding='utf-8')
        for i in tqdm(range(len(ETL))):
            try:
                dtable = Select_datos_sql(
                    NOMINA_SEGURO(str(config("ConsultaSeguros")),
                                  int(ETL.iloc[i]['RUT_TRABAJADOR']))
                )
                df = pd.DataFrame(dtable)
                rut_pens_tran, Rut_Empresa, Dig_Ver_Compania, Monto_Desc_Valor = df.iloc[0, 0]
                logging.info(Monto_Desc_Valor)
            except:
                rut_pens_tran = 0
                Rut_Empresa = 0
                Dig_Ver_Compania = 0
                Monto_Desc_Valor = 0
                if rut_pens_tran == 0 and Rut_Empresa == 0 and Dig_Ver_Compania == 0 and Monto_Desc_Valor == 0:
                    rut_pens_tran = ETL.iloc[i]['RUT_TRABAJADOR']
                    Rut_Empresa = ETL.iloc[i]['RUT_EMPRESA']
                    Dig_Ver_Compania = ETL.iloc[i]['DV_EMPRESA']
                    Monto_Desc_Valor = 0

                else:
                    pass
                pass
            finally:

                listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['RUT_TRABAJADOR'],
                                 'RUT_EMPRESA': Rut_Empresa,
                                 'DV_EMPRESA': Dig_Ver_Compania,
                                 'MONTO': ETL.iloc[i]['MONTO'],
                                 'CRUCE CON NOMINA DE CREDITO': ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'],
                                 'CRUCE CON NOMINA DE SEGUROS': float(Monto_Desc_Valor),
                                 'CRUCE HIPOTECARIO': ' ',
                                 'DIFERENCIAS': ' ',
                                 })

            print(" ", end='\r')
        lETLpaso2 = pd.DataFrame(listaETL)
        lETLpaso2.to_csv("SalidaCVS/ETL02.csv", encoding='utf-8')
    def Facturado_Hipotecario():
        # Consulta Facturado Trabajador Hipotecario
        print("Consulta Facturado Trabajador Hipotecario")
        ETL = pd.read_csv("SalidaCVS/ETL02.csv", encoding='utf-8')
        for i in tqdm(range(len(ETL))):
            try:
                dtable = Select_datos_sql(
                    nomina_hipo_trab_vcto(str(config("ConsultaSeguros")),
                                  int(ETL.iloc[i]['RUT_TRABAJADOR']))
                )
                df = pd.DataFrame(dtable)
                rut_pens_tran, Rut_Empresa, Monto_Desc_Valor = df.iloc[0, 0]
            except:
                rut_pens_tran = 0
                Rut_Empresa = 0
                Monto_Desc_Valor = 0
                if rut_pens_tran == 0 and Rut_Empresa == 0 and Monto_Desc_Valor == 0:
                    rut_pens_tran = ETL.iloc[i]['RUT_TRABAJADOR']
                    Rut_Empresa = ETL.iloc[i]['RUT_EMPRESA']
                    Dig_Ver_Compania = ETL.iloc[i]['DV_EMPRESA']
                    Monto_Desc_Valor = 0
                else:
                    pass
                pass
            finally:
                listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['RUT_TRABAJADOR'],
                                 'RUT_EMPRESA': Rut_Empresa,
                                 'DV_EMPRESA': Dig_Ver_Compania,
                                 'MONTO': ETL.iloc[i]['MONTO'],
                                 'CRUCE CON NOMINA DE CREDITO': ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'],
                                 'CRUCE CON NOMINA DE SEGUROS': ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'],
                                 'CRUCE HIPOTECARIO': float(Monto_Desc_Valor),
                                 'DIFERENCIAS': ' ',
                                 })

            print(" ", end='\r')
        lETLpaso2 = pd.DataFrame(listaETL)
        lETLpaso2.to_csv("SalidaCVS/ETL03.csv", encoding='utf-8')
    def prelar():
        """
        comienza hacer la prelacion con las siguientes operracioes ejemplo=
            SALDO HIPOTECARIO=MONTO-CRUCE HIPOTECARIO
            SALDO CREDITO SOCIAL=SALDO HIPOTECARIO-CRUCE CON NOMINA DE CREDITO
            SALDO SEGURO=SALDO CREDITO SOCIAL-CRUCE CON NOMINA DE SEGUROS
        
        :return: SALDO FINAL
        """
        print("Prelando Previred" )
        ETL = pd.read_csv("SalidaCVS/ETL03.csv", encoding='utf-8' ,index_col=False)


        for i in tqdm(range(len(ETL))):
                rut =ETL.iloc[i]['RUT_TRABAJADOR']
                rut_empresa = ETL.iloc[i]['RUT_EMPRESA']
                dv = ETL.iloc[i]['DV_EMPRESA']
                monto = float(ETL.iloc[i]['MONTO'])
                try:chipotecario = int(ETL.iloc[i]['CRUCE HIPOTECARIO'])
                except:chipotecario = 0
                ccredito= float(ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'])
                try:cseguros = float(ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'])
                except:cseguros = "0"
                if chipotecario == "nan" and chipotecario == " " :chipotecario=0
                if ccredito == "nan" and ccredito == " ": ccredito = 0
                if cseguros == "nan" and cseguros == " ": cseguros = 0
                try:
                     saldohipotecario=0
                     saldocredito=0
                     saldoseguro=0
                except:
                    try:saldohipotecario = 0
                    except: saldohipotecario=0
                    try: saldocredito=0
                    except: saldocredito=0
                    try: saldoseguro=0
                    except: saldoseguro=0
                    pass

                listaETL.append({'RUT_TRABAJADOR': rut,
                                 'RUT_EMPRESA': rut_empresa,
                                 'DV_EMPRESA': dv,
                                 'MONTO': monto,
                                 'CRUCE HIPOTECARIO':float(chipotecario),
                                 'CRUCE CON NOMINA DE CREDITO':float(ccredito),
                                 'CRUCE CON NOMINA DE SEGUROS': float(cseguros),
                                 'SALDO HIPOTECARIO': float(saldohipotecario ),
                                 'SALDO CREDITO': float(saldocredito),
                                 'SALDO SEGURO': float(saldoseguro),
                                 })
                print(" ", end='\r')
        lETLpaso1 = pd.DataFrame(listaETL)
        lETLpaso1.to_excel('SalidaExcel//prelacionPrevired.xlsx', sheet_name='A4', index=False, header=True)
        prelacion = 'SalidaExcel/prelacionPrevired.xlsx'
        workbook = openpyxl.load_workbook(prelacion)
        hoja = workbook["A4"]
        total_registros = hoja.max_row
        logging.info("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
        print("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
        for i in tqdm(range(total_registros)):
            celda=int(i+2)


            try:
                rutPersona = hoja.cell(row=celda, column=1).value
                rutempresa= hoja.cell(row=celda, column=2).value
                dv = hoja.cell(row=celda, column=3).value
            except:
                rutPersona = "-"
                rutempresa= "-"
                dv = "-"
                pass
            if rutPersona is None: rutPersona = ""
            if rutempresa is None: rutempresa = ""
            if dv is None: dv = ""

            try:montos = hoja.cell(row=celda, column=4).value
            except:pass

            # convertir en numero

            hoja.cell(row=celda, column=7).number_format = '0'
            logging.info("ver salida  → "+str(hoja.cell(row=celda, column=7).value))

            if type(hoja.cell(row=celda, column=7).value) == int:
                logging.info(" es numero → "+str(hoja.cell(row=celda, column=7).value))

            else:
                logging.info(" no es numero → " + str(hoja.cell(row=celda, column=7).value))
                hoja.cell(row=celda, column=7).number_format = '0'
                if type(hoja.cell(row=celda, column=7).value) == str:
                    logging.info(" es un texto → " + str(hoja.cell(row=celda, column=7).value))

                    capturamos = str(hoja.cell(row=celda, column=7).value).replace(' ',"")
                    hoja.cell(row=celda, column=7).value = capturamos*1
                    hoja.cell(row=celda, column=7).number_format = '0'

                    #Cuando este vacia la celda .
                    try:
                        if capturamos < 0:
                            hoja.cell(row=celda, column=7).number_format = '0'
                            hoja.cell(row=celda, column=7).value = 0
                    except:

                        if type(capturamos) == int or type(capturamos) == str or capturamos is None or capturamos is not None:
                            logging.info(type(capturamos))
                            validacion = len(hoja.cell(row=celda, column=7).value)
                            if validacion == 0:
                                hoja.cell(row=celda, column=7).value = int(validacion)
                                hoja.cell(row=celda, column=7).number_format = '0'

                else:
                    logging.info("No es un texto → " + str(hoja.cell(row=celda, column=7).value))

            #volvemos asignar las variables
            cruce_hipotecario = hoja.cell(row=celda, column=5).value
            cruce_credito = hoja.cell(row=celda, column=6).value
            cruce_seguros = hoja.cell(row=celda, column=7).value

            if montos is None: montos = 0
            if cruce_hipotecario is None: hoja.cell(row=celda, column=5).value = 0
            if cruce_credito is None: hoja.cell(row=celda, column=6).value = 0
            if cruce_seguros is None:
                hoja.cell(row=celda, column=7).value = 0
            else:
                hoja.cell(row=celda, column=7).value = hoja.cell(row=celda, column=7).value


            try:
                if int(hoja.cell(row=celda, column=5).value) > 0:
                   pass
                else:hoja.cell(row=celda, column=5).value = 0
            except:
                pass
            # calculando los saldos y escribiendolos enxcel
            try:
                saldo_hipotecario = (float(hoja.cell(row=celda, column=4).value)-float(hoja.cell(row=celda, column=5).value))
                hoja.cell(row=celda, column=8).value=(float(hoja.cell(row=celda, column=4).value)-float(hoja.cell(row=celda, column=5).value))
                saldo_hipotecario = hoja.cell(row=celda, column=8).value

            except:saldo_hipotecario =0
            try:
                saldo_credito = float(hoja.cell(row=celda, column=8).value)-float(hoja.cell(row=celda, column=6).value)
                hoja.cell(row=celda, column=9).value = saldo_credito
                saldo_credito = hoja.cell(row=celda, column=9).value
            except:saldo_credito = 0
            try:
                saldo_seguros = float(hoja.cell(row=celda, column=9).value )-float(hoja.cell(row=celda, column=7).value)
                hoja.cell(row=celda, column=10).value = saldo_seguros
                saldo_seguros = hoja.cell(row=celda, column=10).value

            except:saldo_seguros = 0

            #Escribimos el saldo final en el excel
            hoja.cell(row=1, column=11).value= "SALDO FINAL CON TODAS LAS DEDUCCIONES"
            hoja.cell(row=celda, column=11).value = hoja.cell(row=celda, column=10).value

            """print("|-------------Nueva consulta-------------------|")
            print("|rut persona = " + str(rutPersona))
            print("|rut empresa = " + str(rutempresa)+"-"+str(dv))
            print("|montos = "+str(montos))
            print("|cruce_hipotecario = " + str(cruce_hipotecario))
            print("|cruce_credito = " + str(cruce_credito))
            print("|cruce_seguros = " + str(cruce_seguros))
            print("|--------------prelacion-----------------------|")
            print("|Descontamos creditos hipotecarios = "+str(saldo_hipotecario))
            print("|Descontamos saldos de creditos  = " + str(saldo_credito))
            print("|Descontamos saldo seguros  = " + str(saldo_seguros))
            print("|______________________________________________|")
            print("|saldo final = "+str(saldo_seguros))
            print("|______________Consulta finalizada ____________|")"""

            logging.info("|-------------Nueva consulta-------------------|")
            logging.info("|rut persona = " + str(rutPersona))
            logging.info("|rut empresa = " + str(rutempresa)+"-"+str(dv))
            logging.info("|montos = "+str(montos))
            logging.info("|cruce_hipotecario = " + str(cruce_hipotecario))
            logging.info("|cruce_credito = " + str(cruce_credito))
            logging.info("|cruce_seguros = " + str(cruce_seguros))
            logging.info("|--------------prelacion-----------------------|")
            logging.info("|Descontamos creditos hipotecarios = "+str(saldo_hipotecario))
            logging.info("|Descontamos saldos de creditos  = " + str(saldo_credito))
            logging.info("|Descontamos saldo seguros  = " + str(saldo_seguros))
            logging.info("|______________________________________________|")
            logging.info("|saldo final = "+str(saldo_seguros))
            logging.info("|______________Consulta finalizada ____________|")
            print(" ", end='\r')
        workbook.save('SalidaExcel/prelacionPrevired.xlsx')
        workbook.close()

    try : Facturado_Crédito()
    except Exception as e: logging.error(e)
    try : Facturado_Seguros()
    except Exception as e: logging.error(e)
    try :Facturado_Hipotecario()
    except Exception as e: logging.error(e)
    try : prelar()
    except Exception as e: logging.error(e)
    print("prelacion terminada")
    logging.info("prelacion terminada")
def leerCarpetaPl ():
    """
    funcion que lee todos los archivos que comience con PL y los almacena en el acrchivo
    "Data/inputPL.txt"
    :return:
    """
    leerPl=str(config("inputPL"))
    leidos=str(config("outPL"))
    print("Leyendo archivos que inicie  por PL...")
    logging.info("Leyendo archivos que comienze por PL")
    for i in tqdm(range(1)):

        contenido = os.listdir(leerPl)
        for list in contenido:
            try:
                print("Leyendo los registros  →  "+leerPl+list)
                logging.info("Leyendo los registros  →  "+leerPl+list)
                if list.__contains__("PL"):
                        f = open(str(leerPl+list), "r")
                        w = open("Data/inputPL.txt","a")
                        w.write(f.read())
                        f.close()
                        w.close()
                        shutil.copy(leerPl+list, leidos)
                        os.remove(leerPl+list)
            except:
                print("error en lectura de los registros  →  "+leerPl+list)
                logging.error("error en lectura de los registros  →  "+leerPl+list)

    print(" ", end='\r')
def creaciontxt_PLHR():

    """
    Funcion para generar ala salida del PLHR
    :return: Salida PLHR
    """
    try:
        fijo="PLHR"                             # dato fijo
        year = datetime.datetime.now().year     # Obtener el año actual
        year_str = str(year).zfill(4)           # Asegurarse de que sea un valor numérico de 4 dígitos
        mes = datetime.datetime.now().month     # Obtener el mes actual
        mes_str = str(mes).zfill(2)             # Asegurarse de que sea un valor numérico de 2 dígitos
        dia = datetime.datetime.now().day       # Obtener el día actual
        dia_str = str(dia).zfill(2)             # Asegurarse de que sea un valor numérico de 2 dígitos
        Fijolargo6="009990"                     # Dato fijo
        Identificación_Producto="00"            # Dato fijo
        Identificación_fijo= "01000"            # Dato fijo
        Correlativo="001"                       # Correlativo de carga



        #Ubicacion donde se va guardar la PLH
        ubicacion = config("PLHR")

        def plhrHipotecarioTrabjador():
            #credito hipotecario trabjador plhr
            dfPrewired=pd.read_excel("SalidaExcel/prelacionPrevired.xlsx",sheet_name="A4",index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE HIPOTECARIO',
                                             ]]

            dfPrewired = dfPrewired.rename(columns={ 'CRUCE HIPOTECARIO': 'MONTO' })
            Identificación_Producto = "25"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False)
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))

        def plhrnominaCreditotrabajador ():
            # credito credito nomina  trabjador plhr
            dfPrewired = pd.read_excel("SalidaExcel/prelacionPrevired.xlsx", sheet_name="A4", index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE CON NOMINA DE CREDITO',
                                     ]]
            dfPrewired = dfPrewired.rename(columns={'CRUCE CON NOMINA DE CREDITO': 'MONTO'})
            Identificación_Producto = "05"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False )
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))

        def plhrnominasegurostrabajador ():
            # credito sseguros nomina  trabjador plhr
            dfPrewired = pd.read_excel("SalidaExcel/prelacionPrevired.xlsx", sheet_name="A4", index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE CON NOMINA DE SEGUROS',
                                     ]]
            dfPrewired = dfPrewired.rename(columns={'CRUCE CON NOMINA DE SEGUROS': 'MONTO'})
            Identificación_Producto = "15"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False )
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))

        def plhrHipotecarioPensionado():
            #credito hipotecario Pensionado plhr
            dfPrewired=pd.read_excel("SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx",sheet_name="A4",index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE HIPOTECARIO',
                                             ]]

            dfPrewired = dfPrewired.rename(columns={ 'CRUCE HIPOTECARIO': 'MONTO' })
            Identificación_Producto = "26"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False)
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))

        def plhrnominaCreditoPensionado ():
            # credito credito nomina  Pensionado plhr
            dfPrewired = pd.read_excel("SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx", sheet_name="A4", index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE CON NOMINA DE CREDITO',
                                     ]]
            dfPrewired = dfPrewired.rename(columns={'CRUCE CON NOMINA DE CREDITO': 'MONTO'})
            Identificación_Producto = "06"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False )
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))
        def plhrnominasegurosPensionado ():
            # credito sseguros nomina  pensionado plhr
            dfPrewired = pd.read_excel("SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx", sheet_name="A4", index_col=False)
            dfPrewired = dfPrewired[['RUT_TRABAJADOR',
                                     'RUT_EMPRESA',
                                     'DV_EMPRESA',
                                     'CRUCE CON NOMINA DE SEGUROS',
                                     ]]
            dfPrewired = dfPrewired.rename(columns={'CRUCE CON NOMINA DE SEGUROS': 'MONTO'})
            Identificación_Producto = "16"
            # Creacion PLH
            nombre = fijo + year_str + mes_str + year_str + mes_str + dia_str \
                     + Fijolargo6 + Identificación_Producto + Identificación_fijo + Correlativo + ".txt"

            dfPrewired.to_csv(ubicacion + nombre, sep=" ", index=False )
            logging.info("Archivo creado con exito {}".format(nombre))
            print("Archivo creado con exito {}".format(nombre))

        plhrnominasegurostrabajador()
        plhrHipotecarioTrabjador()
        plhrnominaCreditotrabajador()
        plhrnominasegurosPensionado()
        plhrnominaCreditoPensionado()
        plhrHipotecarioPensionado()

        for i in tqdm(range(20)):
            #Apoyo visual para saber el progreso del bot ,puede ser borrado  y no afectara el codigo
            time.sleep(000000000000000.1)
            print(" ", end='\r')
    except Exception as e: logging.error(e)
def prelacionRevicion():
    listaETL = [{}]
    def prelacion():
        """
        comienza hacer la prelacion con las siguientes operracioes ejemplo=
            SALDO HIPOTECARIO=MONTO-CRUCE HIPOTECARIO
            SALDO CREDITO SOCIAL=SALDO HIPOTECARIO-CRUCE CON NOMINA DE CREDITO
            SALDO SEGURO=SALDO CREDITO SOCIAL-CRUCE CON NOMINA DE SEGUROS

        :return: SALDO FINAL
        """
        try:
            print("Prelando Previred verificacion")
            ETL = pd.read_csv("SalidaCVS/ETL03.csv", encoding='utf-8', index_col=False)

            for i in tqdm(range(len(ETL))):
                rut = ETL.iloc[i]['RUT_TRABAJADOR']
                rut_empresa = ETL.iloc[i]['RUT_EMPRESA']
                dv = ETL.iloc[i]['DV_EMPRESA']
                monto = float(ETL.iloc[i]['MONTO'])
                try:
                    chipotecario = int(ETL.iloc[i]['CRUCE HIPOTECARIO'])
                except:
                    chipotecario = 0
                ccredito = float(ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'])
                try:
                    cseguros = float(ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'])
                except:
                    cseguros = "0"
                if chipotecario == "nan" and chipotecario == " ": chipotecario = 0
                if ccredito == "nan" and ccredito == " ": ccredito = 0
                if cseguros == "nan" and cseguros == " ": cseguros = 0
                try:
                    saldohipotecario = 0
                    saldocredito = 0
                    saldoseguro = 0
                except:
                    try:
                        saldohipotecario = 0
                    except:
                        saldohipotecario = 0
                    try:
                        saldocredito = 0
                    except:
                        saldocredito = 0
                    try:
                        saldoseguro = 0
                    except:
                        saldoseguro = 0
                    pass

                listaETL.append({'RUT_TRABAJADOR': rut,
                                 'RUT_EMPRESA': rut_empresa,
                                 'DV_EMPRESA': dv,
                                 'MONTO': monto,
                                 'CRUCE HIPOTECARIO': float(chipotecario),
                                 'CRUCE CON NOMINA DE CREDITO': float(ccredito),
                                 'CRUCE CON NOMINA DE SEGUROS': float(cseguros),
                                 'SALDO HIPOTECARIO': float(saldohipotecario),
                                 'SALDO CREDITO': float(saldocredito),
                                 'SALDO SEGURO': float(saldoseguro),
                                 })
                print(" ", end='\r')
            lETLpaso1 = pd.DataFrame(listaETL)
            lETLpaso1.to_excel('SalidaExcel//prelacionPrevired.xlsx', sheet_name='A4', index=False, header=True)
            prelacion = 'SalidaExcel/prelacionPrevired.xlsx'
            workbook = openpyxl.load_workbook(prelacion)
            hoja = workbook["A4"]
            total_registros = hoja.max_row
            logging.info("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
            print("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
            for i in tqdm(range(total_registros)):
                celda = int(i + 2)

                try:
                    rutPersona = hoja.cell(row=celda, column=1).value
                    rutempresa = hoja.cell(row=celda, column=2).value
                    dv = hoja.cell(row=celda, column=3).value
                except:
                    rutPersona = "-"
                    rutempresa = "-"
                    dv = "-"
                    pass
                if rutPersona is None: rutPersona = ""
                if rutempresa is None: rutempresa = ""
                if dv is None: dv = ""

                try:
                    montos = hoja.cell(row=celda, column=4).value
                except:
                    pass

                # convertir en numero

                hoja.cell(row=celda, column=7).number_format = '0'
                logging.info("ver salida  → " + str(hoja.cell(row=celda, column=7).value))

                if type(hoja.cell(row=celda, column=7).value) == int:
                    logging.info(" es numero → " + str(hoja.cell(row=celda, column=7).value))

                else:
                    logging.info(" no es numero → " + str(hoja.cell(row=celda, column=7).value))
                    hoja.cell(row=celda, column=7).number_format = '0'
                    if type(hoja.cell(row=celda, column=7).value) == str:
                        logging.info(" es un texto → " + str(hoja.cell(row=celda, column=7).value))

                        capturamos = str(hoja.cell(row=celda, column=7).value).replace(' ', "")
                        hoja.cell(row=celda, column=7).value = capturamos * 1
                        hoja.cell(row=celda, column=7).number_format = '0'

                        # Cuando este vacia la celda .
                        try:
                            if capturamos < 0:
                                hoja.cell(row=celda, column=7).number_format = '0'
                                hoja.cell(row=celda, column=7).value = 0
                        except:

                            if type(capturamos) == int or type(
                                    capturamos) == str or capturamos is None or capturamos is not None:
                                logging.info(type(capturamos))
                                validacion = len(hoja.cell(row=celda, column=7).value)
                                if validacion == 0:
                                    hoja.cell(row=celda, column=7).value = int(validacion)
                                    hoja.cell(row=celda, column=7).number_format = '0'

                    else:
                        logging.info("No es un texto → " + str(hoja.cell(row=celda, column=7).value))

                # volvemos asignar las variables
                cruce_hipotecario = hoja.cell(row=celda, column=5).value
                cruce_credito = hoja.cell(row=celda, column=6).value
                cruce_seguros = hoja.cell(row=celda, column=7).value

                if montos is None: montos = 0
                if cruce_hipotecario is None: hoja.cell(row=celda, column=5).value = 0
                if cruce_credito is None: hoja.cell(row=celda, column=6).value = 0
                if cruce_seguros is None:
                    hoja.cell(row=celda, column=7).value = 0
                else:
                    hoja.cell(row=celda, column=7).value = hoja.cell(row=celda, column=7).value

                try:
                    if int(hoja.cell(row=celda, column=5).value) > 0:
                        pass
                    else:
                        hoja.cell(row=celda, column=5).value = 0
                except:
                    pass
                # calculando los saldos y escribiendolos enxcel
                try:
                    saldo_hipotecario = (float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                    hoja.cell(row=celda, column=8).value = (float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                    saldo_hipotecario = hoja.cell(row=celda, column=8).value

                except UnboundLocalError:
                    print("revision " + str(saldo_hipotecario))
                except:
                    saldo_hipotecario = 0

                try:
                    saldo_credito = float(hoja.cell(row=celda, column=8).value) - float(
                        hoja.cell(row=celda, column=6).value)
                    hoja.cell(row=celda, column=9).value = saldo_credito
                    saldo_credito = hoja.cell(row=celda, column=9).value
                except:
                    saldo_credito = 0
                try:
                    saldo_seguros = float(hoja.cell(row=celda, column=9).value) - float(
                        hoja.cell(row=celda, column=7).value)
                    hoja.cell(row=celda, column=10).value = saldo_seguros
                    saldo_seguros = hoja.cell(row=celda, column=10).value

                except:
                    saldo_seguros = 0

                # Escribimos el saldo final en el excel
                hoja.cell(row=1, column=11).value = "SALDO FINAL CON TODAS LAS DEDUCCIONES"
                hoja.cell(row=celda, column=11).value = hoja.cell(row=celda, column=10).value

                """print("|-------------Nueva consulta-------------------|")
                print("|rut persona = " + str(rutPersona))
                print("|rut empresa = " + str(rutempresa)+"-"+str(dv))
                print("|montos = "+str(montos))
                print("|cruce_hipotecario = " + str(cruce_hipotecario))
                print("|cruce_credito = " + str(cruce_credito))
                print("|cruce_seguros = " + str(cruce_seguros))
                print("|--------------prelacion-----------------------|")
                print("|Descontamos creditos hipotecarios = "+str(saldo_hipotecario))
                print("|Descontamos saldos de creditos  = " + str(saldo_credito))
                print("|Descontamos saldo seguros  = " + str(saldo_seguros))
                print("|______________________________________________|")
                print("|saldo final = "+str(saldo_seguros))
                print("|______________Consulta finalizada ____________|")"""

                logging.info("|-------------Nueva consulta-------------------|")
                logging.info("|rut persona = " + str(rutPersona))
                logging.info("|rut empresa = " + str(rutempresa) + "-" + str(dv))
                logging.info("|montos = " + str(montos))
                logging.info("|cruce_hipotecario = " + str(cruce_hipotecario))
                logging.info("|cruce_credito = " + str(cruce_credito))
                logging.info("|cruce_seguros = " + str(cruce_seguros))
                logging.info("|--------------prelacion-----------------------|")
                logging.info("|Descontamos creditos hipotecarios = " + str(saldo_hipotecario))
                logging.info("|Descontamos saldos de creditos  = " + str(saldo_credito))
                logging.info("|Descontamos saldo seguros  = " + str(saldo_seguros))
                logging.info("|______________________________________________|")
                logging.info("|saldo final = " + str(saldo_seguros))
                logging.info("|______________Consulta finalizada ____________|")
                print(" ", end='\r')
            workbook.save('SalidaExcel/prelacionPrevired.xlsx')
            workbook.close()
        except Exception as e: logging.error(e)

        # CONFIRMACION
        prelacion()
        print("validacion prelacion terminada")
        logging.info("validacion prelacion terminada")
def HE():
    """      modelo que extrae todos los datos de archivo txt que inicia con "HE":
            :return: lectura de todos los SERVIPAG
    """
    try:
        path = "Data/inputHE.txt"
        logging.info("Abriendo Txt que inicia con HERO")
        f = open(path)
        f = f.readlines()
        for x in f:
            #oficina=x[0:3]
            #id_papel = x[3:13]
            #id_ticket = x[13:23]
            #Identificador=x[23:33]
            #Numero_Documento = x[33:48]
            Rut = x[48:58]
            #interes = x[58:66]
            #Cobranza = x[66:74]
            #Cuota = x[74:77]
            #Fecha_Vencimiento = x[74:82]
            #Monto_Minimo = x[82:90]
            Monto = x[90:98]
            #MediodePago = x[98:100]
            #n_serie=x[100:112]
            #Banco=x[112:115]
            #plaza_docto=x[115:119]
            #cuenta_cte=x[119:131]
            #fechapago=x[131:157]
            #hora = x[157:163]
            #fechacontable=x[163:171]
            #referncia = x[171:191]

            try:
                servipag.append({
                            'RUT_TRABAJADOR': Rut,
                            'RUT_EMPRESA': 0,
                            'DV_EMPRESA': 0,
                            'MONTO': float(Monto),
                    })
            except ValueError as e:
                print(e)
                servipag.append({
                            'RUT_TRABAJADOR': 0,
                            'RUT_EMPRESA': 0,
                            'DV_EMPRESA': 0,
                            'MONTO': 0,})

        df = pd.DataFrame(servipag)
        df = df[['RUT_TRABAJADOR',
                 'RUT_EMPRESA',
                 'DV_EMPRESA',
                 'MONTO'
                 ]]
        df.to_csv("SalidaCVS/ETLHE04.csv", index=False)
        print("lectura de HE's terminada...")
        for i in tqdm(range(20)):
            # Apoyo visual pasa saber el progreso del bot ser borrado  y no afectara el codigo
            time.sleep(000000000000000.1)
            print(" ", end='\r')

    except:
        pass
def leerCarpetaHE ():
    """
    funcion que lee todos los archivos que comience con PL y los almacena en el acrchivo
    "Data/inputHE.txt"
    :return:
    """
    leerPl=str(config("inputPL"))
    leidos=str(config("outPL"))
    print("Leyendo archivos que inicie  por PL...")
    logging.info("Leyendo archivos que comienze por PL")
    for i in tqdm(range(1)):

        contenido = os.listdir(leerPl)
        for list in contenido:
            try:
                print("Leyendo los registros  →  "+leerPl+list)
                logging.info("Leyendo los registros  →  "+leerPl+list)
                if list.__contains__("HE"):
                        f = open(str(leerPl+list), "r")
                        w = open("Data/inputHE.txt","a")
                        w.write(f.read())
                        f.close()
                        w.close()
                        shutil.copy(leerPl+list, leidos)
                        os.remove(leerPl+list)
            except:
                print("error en lectura de los registros  →  "+leerPl+list)
                logging.error("error en lectura de los registros  →  "+leerPl+list)
def ETLA4_PENSIONADO():
    """
    ETL que hace las consultas en:
    1. Consulta Facturado pensionado Crédito.
    2. Consulta Facturado pensionado Seguros
    3. Consulta Facturado pensionado Hipotecario
    :return: Prelacion previred pensionado.
    """

    listaETL = [{}]

    def Facturado_Crédito_PENSIONADO():
        print("Consulta Facturado pensionado  Crédito")
        ETL = pd.read_csv("SalidaCVS/index.csv")

        # Consulta Facturado pensinado Crédito
        for i in tqdm(range(len(ETL))):

            try:
                dtable = Select_datos_sql(
                    NOMINA_COBRO_pens_VCTO_(str(config("Consulta")),
                                                 ETL.iloc[i]['Rut_del_trabajador'])
                )
                df = pd.DataFrame(dtable)
                rut_pens_tran, Rut_Empresa, Dig_Ver_Compania, Monto_Desc_Valor = df.iloc[0, 0]

            except:
                rut_pens_tran = 0
                Rut_Empresa = 0
                Dig_Ver_Compania = 0
                Monto_Desc_Valor = 0
                if rut_pens_tran == 0 and Rut_Empresa == 0 and Dig_Ver_Compania == 0 and Monto_Desc_Valor == 0:
                    rut_pens_tran = ""
                    Rut_Empresa = ""
                    Dig_Ver_Compania = ""
                    Monto_Desc_Valor = 0
                pass
            finally:

                listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['Rut_del_trabajador'],
                                 'RUT_EMPRESA': Rut_Empresa,
                                 'DV_EMPRESA': Dig_Ver_Compania,
                                 'MONTO': ETL.iloc[i]['Productos_financieros'],
                                 'CRUCE CON NOMINA DE CREDITO': Monto_Desc_Valor,
                                 'CRUCE CON NOMINA DE SEGUROS': ' ',
                                 'CRUCE HIPOTECARIO': ' ',
                                 'DIFERENCIAS': ' ',
                                 })

            print(" ", end='\r')  # Consulta Facturado Trabajador Crédito #Consulta Facturado Trabajador Crédito
        lETLpaso1 = pd.DataFrame(listaETL)
        lETLpaso1.to_csv("SalidaCVS/ETL01P.csv")

    def Facturado_Seguros_PENSIONADO():
        # Consulta Facturado pensionado Seguros
        print("Consulta Facturado pensionado Seguros")
        ETL = pd.read_csv("SalidaCVS/ETL01P.csv", encoding='utf-8')
        for i in tqdm(range(len(ETL))):
            try:
                dtable = Select_datos_sql(
                    NOMINA_SEGURO_PENSIONADO(str(config("ConsultaSeguros")),
                                  int(ETL.iloc[i]['RUT_TRABAJADOR']))
                )
                df = pd.DataFrame(dtable)
                rut_pens_tran, Rut_Empresa, Dig_Ver_Compania, Monto_Desc_Valor = df.iloc[0, 0]
                logging.info(Monto_Desc_Valor)
            except:
                rut_pens_tran = 0
                Rut_Empresa = 0
                Dig_Ver_Compania = 0
                Monto_Desc_Valor = 0
                if rut_pens_tran == 0 and Rut_Empresa == 0 and Dig_Ver_Compania == 0 and Monto_Desc_Valor == 0:
                    rut_pens_tran = ETL.iloc[i]['RUT_TRABAJADOR']
                    Rut_Empresa = ETL.iloc[i]['RUT_EMPRESA']
                    Dig_Ver_Compania = ETL.iloc[i]['DV_EMPRESA']
                    Monto_Desc_Valor = 0

                else:
                    pass
                pass
            finally:

                listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['RUT_TRABAJADOR'],
                                 'RUT_EMPRESA': Rut_Empresa,
                                 'DV_EMPRESA': Dig_Ver_Compania,
                                 'MONTO': ETL.iloc[i]['MONTO'],
                                 'CRUCE CON NOMINA DE CREDITO': ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'],
                                 'CRUCE CON NOMINA DE SEGUROS': float(Monto_Desc_Valor),
                                 'CRUCE HIPOTECARIO': ' ',
                                 'DIFERENCIAS': ' ',
                                 })

            print(" ", end='\r')
        lETLpaso2 = pd.DataFrame(listaETL)
        lETLpaso2.to_csv("SalidaCVS/ETL02P.csv", encoding='utf-8')

    def Facturado_Hipotecario_PENSIONADO():
        # Consulta Facturado Pensionado Hipotecario
        print("Consulta Facturado Pensionado Hipotecario")
        ETL = pd.read_csv("SalidaCVS/ETL02P.csv", encoding='utf-8')
        for i in tqdm(range(len(ETL))):
            try:
                dtable = Select_datos_sql(
                    nomina_hipo_pens_vcto_(str(config("ConsultaSeguros")),
                                          int(ETL.iloc[i]['RUT_TRABAJADOR']))
                )
                df = pd.DataFrame(dtable)
                rut_pens_tran, Rut_Empresa, Monto_Desc_Valor = df.iloc[0, 0]
            except:
                rut_pens_tran = 0
                Rut_Empresa = 0
                Monto_Desc_Valor = 0
                if rut_pens_tran == 0 and Rut_Empresa == 0 and Monto_Desc_Valor == 0:
                    rut_pens_tran = ETL.iloc[i]['RUT_TRABAJADOR']
                    Rut_Empresa = ETL.iloc[i]['RUT_EMPRESA']
                    Dig_Ver_Compania = ETL.iloc[i]['DV_EMPRESA']
                    Monto_Desc_Valor = 0
                else:
                    pass
                pass
            finally:
                listaETL.append({'RUT_TRABAJADOR': ETL.iloc[i]['RUT_TRABAJADOR'],
                                 'RUT_EMPRESA': Rut_Empresa,
                                 'DV_EMPRESA': Dig_Ver_Compania,
                                 'MONTO': ETL.iloc[i]['MONTO'],
                                 'CRUCE CON NOMINA DE CREDITO': ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'],
                                 'CRUCE CON NOMINA DE SEGUROS': ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'],
                                 'CRUCE HIPOTECARIO': float(Monto_Desc_Valor),
                                 'DIFERENCIAS': ' ',
                                 })

            print(" ", end='\r')
        lETLpaso2 = pd.DataFrame(listaETL)
        lETLpaso2.to_csv("SalidaCVS/ETL03P.csv", encoding='utf-8')

    def prelar_PENSIONADO():
        """
        comienza hacer la prelacion con las siguientes operracioes ejemplo=
            SALDO HIPOTECARIO=MONTO-CRUCE HIPOTECARIO
            SALDO CREDITO SOCIAL=SALDO HIPOTECARIO-CRUCE CON NOMINA DE CREDITO
            SALDO SEGURO=SALDO CREDITO SOCIAL-CRUCE CON NOMINA DE SEGUROS

        :return: SALDO FINAL
        """
        print("Prelando Previred Pensionado")
        ETL = pd.read_csv("SalidaCVS/ETL03P.csv", encoding='utf-8', index_col=False)

        for i in tqdm(range(len(ETL))):
            rut = ETL.iloc[i]['RUT_TRABAJADOR']
            rut_empresa = ETL.iloc[i]['RUT_EMPRESA']
            dv = ETL.iloc[i]['DV_EMPRESA']
            monto = float(ETL.iloc[i]['MONTO'])
            try:
                chipotecario = int(ETL.iloc[i]['CRUCE HIPOTECARIO'])
            except:
                chipotecario = 0
            ccredito = float(ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'])
            try:
                cseguros = float(ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'])
            except:
                cseguros = "0"
            if chipotecario == "nan" and chipotecario == " ": chipotecario = 0
            if ccredito == "nan" and ccredito == " ": ccredito = 0
            if cseguros == "nan" and cseguros == " ": cseguros = 0
            try:
                saldohipotecario = 0
                saldocredito = 0
                saldoseguro = 0
            except:
                try:
                    saldohipotecario = 0
                except:
                    saldohipotecario = 0
                try:
                    saldocredito = 0
                except:
                    saldocredito = 0
                try:
                    saldoseguro = 0
                except:
                    saldoseguro = 0
                pass

            listaETL.append({'RUT_TRABAJADOR': rut,
                             'RUT_EMPRESA': rut_empresa,
                             'DV_EMPRESA': dv,
                             'MONTO': monto,
                             'CRUCE HIPOTECARIO': float(chipotecario),
                             'CRUCE CON NOMINA DE CREDITO': float(ccredito),
                             'CRUCE CON NOMINA DE SEGUROS': float(cseguros),
                             'SALDO HIPOTECARIO': float(saldohipotecario),
                             'SALDO CREDITO': float(saldocredito),
                             'SALDO SEGURO': float(saldoseguro),
                             })
            print(" ", end='\r')
        lETLpaso1 = pd.DataFrame(listaETL)
        lETLpaso1.to_excel('SalidaExcel//prelacionPrevired_PENSIONADOS.xlsx', sheet_name='A4', index=False, header=True)
        prelacion = 'SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx'
        workbook = openpyxl.load_workbook(prelacion)
        hoja = workbook["A4"]
        total_registros = hoja.max_row
        logging.info("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
        print("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
        for i in tqdm(range(total_registros)):
            celda = int(i + 2)

            try:
                rutPersona = hoja.cell(row=celda, column=1).value
                rutempresa = hoja.cell(row=celda, column=2).value
                dv = hoja.cell(row=celda, column=3).value
            except:
                rutPersona = "-"
                rutempresa = "-"
                dv = "-"
                pass
            if rutPersona is None: rutPersona = ""
            if rutempresa is None: rutempresa = ""
            if dv is None: dv = ""

            try:
                montos = hoja.cell(row=celda, column=4).value
            except:
                pass

            # convertir en numero

            hoja.cell(row=celda, column=7).number_format = '0'
            logging.info("ver salida  → " + str(hoja.cell(row=celda, column=7).value))

            if type(hoja.cell(row=celda, column=7).value) == int:
                logging.info(" es numero → " + str(hoja.cell(row=celda, column=7).value))

            else:
                logging.info(" no es numero → " + str(hoja.cell(row=celda, column=7).value))
                hoja.cell(row=celda, column=7).number_format = '0'
                if type(hoja.cell(row=celda, column=7).value) == str:
                    logging.info(" es un texto → " + str(hoja.cell(row=celda, column=7).value))

                    capturamos = str(hoja.cell(row=celda, column=7).value).replace(' ', "")
                    hoja.cell(row=celda, column=7).value = capturamos * 1
                    hoja.cell(row=celda, column=7).number_format = '0'

                    # Cuando este vacia la celda .
                    try:
                        if capturamos < 0:
                            hoja.cell(row=celda, column=7).number_format = '0'
                            hoja.cell(row=celda, column=7).value = 0
                    except:

                        if type(capturamos) == int or type(
                                capturamos) == str or capturamos is None or capturamos is not None:
                            logging.info(type(capturamos))
                            validacion = len(hoja.cell(row=celda, column=7).value)
                            if validacion == 0:
                                hoja.cell(row=celda, column=7).value = int(validacion)
                                hoja.cell(row=celda, column=7).number_format = '0'

                else:
                    logging.info("No es un texto → " + str(hoja.cell(row=celda, column=7).value))

            # volvemos asignar las variables
            cruce_hipotecario = hoja.cell(row=celda, column=5).value
            cruce_credito = hoja.cell(row=celda, column=6).value
            cruce_seguros = hoja.cell(row=celda, column=7).value

            if montos is None: montos = 0
            if cruce_hipotecario is None: hoja.cell(row=celda, column=5).value = 0
            if cruce_credito is None: hoja.cell(row=celda, column=6).value = 0
            if cruce_seguros is None:
                hoja.cell(row=celda, column=7).value = 0
            else:
                hoja.cell(row=celda, column=7).value = hoja.cell(row=celda, column=7).value

            try:
                if int(hoja.cell(row=celda, column=5).value) > 0:
                    pass
                else:
                    hoja.cell(row=celda, column=5).value = 0
            except:
                pass
            # calculando los saldos y escribiendolos enxcel
            try:
                saldo_hipotecario = (
                            float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                hoja.cell(row=celda, column=8).value = (
                            float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                saldo_hipotecario = hoja.cell(row=celda, column=8).value

            except:
                saldo_hipotecario = 0
            try:
                saldo_credito = float(hoja.cell(row=celda, column=8).value) - float(
                    hoja.cell(row=celda, column=6).value)
                hoja.cell(row=celda, column=9).value = saldo_credito
                saldo_credito = hoja.cell(row=celda, column=9).value
            except:
                saldo_credito = 0
            try:
                saldo_seguros = float(hoja.cell(row=celda, column=9).value) - float(
                    hoja.cell(row=celda, column=7).value)
                hoja.cell(row=celda, column=10).value = saldo_seguros
                saldo_seguros = hoja.cell(row=celda, column=10).value

            except:
                saldo_seguros = 0

            # Escribimos el saldo final en el excel
            hoja.cell(row=1, column=11).value = "SALDO FINAL CON TODAS LAS DEDUCCIONES"
            hoja.cell(row=celda, column=11).value = hoja.cell(row=celda, column=10).value

            """print("|-------------Nueva consulta-------------------|")
            print("|rut persona = " + str(rutPersona))
            print("|rut empresa = " + str(rutempresa)+"-"+str(dv))
            print("|montos = "+str(montos))
            print("|cruce_hipotecario = " + str(cruce_hipotecario))
            print("|cruce_credito = " + str(cruce_credito))
            print("|cruce_seguros = " + str(cruce_seguros))
            print("|--------------prelacion-----------------------|")
            print("|Descontamos creditos hipotecarios = "+str(saldo_hipotecario))
            print("|Descontamos saldos de creditos  = " + str(saldo_credito))
            print("|Descontamos saldo seguros  = " + str(saldo_seguros))
            print("|______________________________________________|")
            print("|saldo final = "+str(saldo_seguros))
            print("|______________Consulta finalizada ____________|")"""

            logging.info("|-------------Nueva consulta-------------------|")
            logging.info("|rut persona = " + str(rutPersona))
            logging.info("|rut empresa = " + str(rutempresa) + "-" + str(dv))
            logging.info("|montos = " + str(montos))
            logging.info("|cruce_hipotecario = " + str(cruce_hipotecario))
            logging.info("|cruce_credito = " + str(cruce_credito))
            logging.info("|cruce_seguros = " + str(cruce_seguros))
            logging.info("|--------------prelacion-----------------------|")
            logging.info("|Descontamos creditos hipotecarios = " + str(saldo_hipotecario))
            logging.info("|Descontamos saldos de creditos  = " + str(saldo_credito))
            logging.info("|Descontamos saldo seguros  = " + str(saldo_seguros))
            logging.info("|______________________________________________|")
            logging.info("|saldo final = " + str(saldo_seguros))
            logging.info("|______________Consulta finalizada ____________|")
            print(" ", end='\r')
        workbook.save('SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx')
        workbook.close()

    try:
        Facturado_Crédito_PENSIONADO()
    except Exception as e:
        logging.error(e)
    try:
        Facturado_Seguros_PENSIONADO()
    except Exception as e:
        logging.error(e)
    try:
        Facturado_Hipotecario_PENSIONADO()
    except Exception as e:
        logging.error(e)
    try:
        prelar_PENSIONADO()
    except Exception as e:
        logging.error(e)
    print("prelacion Pensionados terminada")
    logging.info("prelacion Pensionados terminada")
def prelacionRevicion_PENSIONADO():
    try:
        listaETL = [{}]
        def prelacion():
            """
            comienza hacer la prelacion con las siguientes operracioes ejemplo=
                SALDO HIPOTECARIO=MONTO-CRUCE HIPOTECARIO
                SALDO CREDITO SOCIAL=SALDO HIPOTECARIO-CRUCE CON NOMINA DE CREDITO
                SALDO SEGURO=SALDO CREDITO SOCIAL-CRUCE CON NOMINA DE SEGUROS

            :return: SALDO FINAL
            """
            print("Prelando Previred verificacion")
            ETL = pd.read_csv("SalidaCVS/ETL03P.csv", encoding='utf-8', index_col=False)

            for i in tqdm(range(len(ETL))):
                rut = ETL.iloc[i]['RUT_TRABAJADOR']
                rut_empresa = ETL.iloc[i]['RUT_EMPRESA']
                dv = ETL.iloc[i]['DV_EMPRESA']
                monto = float(ETL.iloc[i]['MONTO'])
                try:
                    chipotecario = int(ETL.iloc[i]['CRUCE HIPOTECARIO'])
                except:
                    chipotecario = 0
                ccredito = float(ETL.iloc[i]['CRUCE CON NOMINA DE CREDITO'])
                try:
                    cseguros = float(ETL.iloc[i]['CRUCE CON NOMINA DE SEGUROS'])
                except:
                    cseguros = "0"
                if chipotecario == "nan" and chipotecario == " ": chipotecario = 0
                if ccredito == "nan" and ccredito == " ": ccredito = 0
                if cseguros == "nan" and cseguros == " ": cseguros = 0
                try:
                    saldohipotecario = 0
                    saldocredito = 0
                    saldoseguro = 0
                except:
                    try:
                        saldohipotecario = 0
                    except:
                        saldohipotecario = 0
                    try:
                        saldocredito = 0
                    except:
                        saldocredito = 0
                    try:
                        saldoseguro = 0
                    except:
                        saldoseguro = 0
                    pass

                listaETL.append({'RUT_TRABAJADOR': rut,
                                 'RUT_EMPRESA': rut_empresa,
                                 'DV_EMPRESA': dv,
                                 'MONTO': monto,
                                 'CRUCE HIPOTECARIO': float(chipotecario),
                                 'CRUCE CON NOMINA DE CREDITO': float(ccredito),
                                 'CRUCE CON NOMINA DE SEGUROS': float(cseguros),
                                 'SALDO HIPOTECARIO': float(saldohipotecario),
                                 'SALDO CREDITO': float(saldocredito),
                                 'SALDO SEGURO': float(saldoseguro),
                                 })
                print(" ", end='\r')
            lETLpaso1 = pd.DataFrame(listaETL)
            lETLpaso1.to_excel('SalidaExcel//prelacionPrevired_PENSIONADOS.xlsx', sheet_name='A4', index=False, header=True)
            prelacion = 'SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx'
            workbook = openpyxl.load_workbook(prelacion)
            hoja = workbook["A4"]
            total_registros = hoja.max_row
            logging.info("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
            print("Comparando los montos Vs saldos hipotecarios , creditos y seguros ...")
            for i in tqdm(range(total_registros)):
                celda = int(i + 2)

                try:
                    rutPersona = hoja.cell(row=celda, column=1).value
                    rutempresa = hoja.cell(row=celda, column=2).value
                    dv = hoja.cell(row=celda, column=3).value
                except:
                    rutPersona = "-"
                    rutempresa = "-"
                    dv = "-"
                    pass
                if rutPersona is None: rutPersona = ""
                if rutempresa is None: rutempresa = ""
                if dv is None: dv = ""

                try:
                    montos = hoja.cell(row=celda, column=4).value
                except:
                    pass

                # convertir en numero

                hoja.cell(row=celda, column=7).number_format = '0'
                logging.info("ver salida  → " + str(hoja.cell(row=celda, column=7).value))

                if type(hoja.cell(row=celda, column=7).value) == int:
                    logging.info(" es numero → " + str(hoja.cell(row=celda, column=7).value))

                else:
                    logging.info(" no es numero → " + str(hoja.cell(row=celda, column=7).value))
                    hoja.cell(row=celda, column=7).number_format = '0'
                    if type(hoja.cell(row=celda, column=7).value) == str:
                        logging.info(" es un texto → " + str(hoja.cell(row=celda, column=7).value))

                        capturamos = str(hoja.cell(row=celda, column=7).value).replace(' ', "")
                        hoja.cell(row=celda, column=7).value = capturamos * 1
                        hoja.cell(row=celda, column=7).number_format = '0'

                        # Cuando este vacia la celda .
                        try:
                            if capturamos < 0:
                                hoja.cell(row=celda, column=7).number_format = '0'
                                hoja.cell(row=celda, column=7).value = 0
                        except:

                            if type(capturamos) == int or type(
                                    capturamos) == str or capturamos is None or capturamos is not None:
                                logging.info(type(capturamos))
                                validacion = len(hoja.cell(row=celda, column=7).value)
                                if validacion == 0:
                                    hoja.cell(row=celda, column=7).value = int(validacion)
                                    hoja.cell(row=celda, column=7).number_format = '0'

                    else:
                        logging.info("No es un texto → " + str(hoja.cell(row=celda, column=7).value))

                # volvemos asignar las variables
                cruce_hipotecario = hoja.cell(row=celda, column=5).value
                cruce_credito = hoja.cell(row=celda, column=6).value
                cruce_seguros = hoja.cell(row=celda, column=7).value

                if montos is None: montos = 0
                if cruce_hipotecario is None: hoja.cell(row=celda, column=5).value = 0
                if cruce_credito is None: hoja.cell(row=celda, column=6).value = 0
                if cruce_seguros is None:
                    hoja.cell(row=celda, column=7).value = 0
                else:
                    hoja.cell(row=celda, column=7).value = hoja.cell(row=celda, column=7).value

                try:
                    if int(hoja.cell(row=celda, column=5).value) > 0:
                        pass
                    else:
                        hoja.cell(row=celda, column=5).value = 0
                except:
                    pass
                # calculando los saldos y escribiendolos enxcel
                try:
                    saldo_hipotecario = (float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                    hoja.cell(row=celda, column=8).value = (float(hoja.cell(row=celda, column=4).value) - float(hoja.cell(row=celda, column=5).value))
                    saldo_hipotecario = hoja.cell(row=celda, column=8).value

                except UnboundLocalError:
                    print("revision " + str(saldo_hipotecario))
                except:
                    saldo_hipotecario = 0

                try:
                    saldo_credito = float(hoja.cell(row=celda, column=8).value) - float(
                        hoja.cell(row=celda, column=6).value)
                    hoja.cell(row=celda, column=9).value = saldo_credito
                    saldo_credito = hoja.cell(row=celda, column=9).value
                except:
                    saldo_credito = 0
                try:
                    saldo_seguros = float(hoja.cell(row=celda, column=9).value) - float(
                        hoja.cell(row=celda, column=7).value)
                    hoja.cell(row=celda, column=10).value = saldo_seguros
                    saldo_seguros = hoja.cell(row=celda, column=10).value

                except:
                    saldo_seguros = 0

                # Escribimos el saldo final en el excel
                hoja.cell(row=1, column=11).value = "SALDO FINAL CON TODAS LAS DEDUCCIONES"
                hoja.cell(row=celda, column=11).value = hoja.cell(row=celda, column=10).value

                """print("|-------------Nueva consulta-------------------|")
                print("|rut persona = " + str(rutPersona))
                print("|rut empresa = " + str(rutempresa)+"-"+str(dv))
                print("|montos = "+str(montos))
                print("|cruce_hipotecario = " + str(cruce_hipotecario))
                print("|cruce_credito = " + str(cruce_credito))
                print("|cruce_seguros = " + str(cruce_seguros))
                print("|--------------prelacion-----------------------|")
                print("|Descontamos creditos hipotecarios = "+str(saldo_hipotecario))
                print("|Descontamos saldos de creditos  = " + str(saldo_credito))
                print("|Descontamos saldo seguros  = " + str(saldo_seguros))
                print("|______________________________________________|")
                print("|saldo final = "+str(saldo_seguros))
                print("|______________Consulta finalizada ____________|")"""

                logging.info("|-------------Nueva consulta-------------------|")
                logging.info("|rut persona = " + str(rutPersona))
                logging.info("|rut empresa = " + str(rutempresa) + "-" + str(dv))
                logging.info("|montos = " + str(montos))
                logging.info("|cruce_hipotecario = " + str(cruce_hipotecario))
                logging.info("|cruce_credito = " + str(cruce_credito))
                logging.info("|cruce_seguros = " + str(cruce_seguros))
                logging.info("|--------------prelacion-----------------------|")
                logging.info("|Descontamos creditos hipotecarios = " + str(saldo_hipotecario))
                logging.info("|Descontamos saldos de creditos  = " + str(saldo_credito))
                logging.info("|Descontamos saldo seguros  = " + str(saldo_seguros))
                logging.info("|______________________________________________|")
                logging.info("|saldo final = " + str(saldo_seguros))
                logging.info("|______________Consulta finalizada ____________|")
                print(" ", end='\r')
            workbook.save('SalidaExcel/prelacionPrevired_PENSIONADOS.xlsx')
            workbook.close()


        # CONFIRMACION
        prelacion()
        print("validacion prelacion terminada")
        logging.info("validacion prelacion terminada")
    except Exception as e: logging.error(e)
def UnionPLyHE():
    try:
        Df=pd.read_csv("SalidaCVS/index.csv")
        for i in range(len(Df)):

            index.append({'Rut_del_trabajador':Df.iloc[i]['Rut_del_trabajador'],
                        'rut':Df.iloc[i]['rut'],
                        'Monto_de_Asignación_Familiar':Df.iloc[i]['Monto_de_Asignación_Familiar'],
                        'Monto_de_Creditos_Personales':Df.iloc[i]['Monto_de_Creditos_Personales'],
                        'Monto_de_Convenios_Dentales':Df.iloc[i]['Monto_de_Convenios_Dentales'],
                        'Monto_Leasing':Df.iloc[i]['Monto_Leasing'],
                        'Montos_de_Seguros_de_Vidas':Df.iloc[i]['Montos_de_Seguros_de_Vidas'],
                        'Aporte_1': Df.iloc[i]['Aporte_1'],
                        'Aporte_Adicional':Df.iloc[i]['Aporte_Adicional'],
                        'Otros_CCAF':Df.iloc[i]['Otros_CCAF'],
                        'Monto_Asignación_Familiar_Retroactiva':Df.iloc[i]['Monto_Asignación_Familiar_Retroactiva'],
                        'Monto_Reintegros_Asignación_Familiar':Df.iloc[i]['Monto_Reintegros_Asignación_Familiar'],
                        'Monto_6_de_trabajador_no_afiliado_a_Isapre':Df.iloc[i]['Monto_6_de_trabajador_no_afiliado_a_Isapre'],
                        'Productos_financieros':Df.iloc[i]['Productos_financieros'],
                        'afa':Df.iloc[i]['afa'],
                        })
        DfHE = pd.read_csv("SalidaCVS/ETLHE04.csv")

        for i in range(len(DfHE)):

            index.append({'Rut_del_trabajador':DfHE.iloc[i]['RUT_TRABAJADOR'],
                        'rut':0,
                        'Monto_de_Asignación_Familiar':0,
                        'Monto_de_Creditos_Personales':0,
                        'Monto_de_Convenios_Dentales':0,
                        'Monto_Leasing':0,
                        'Montos_de_Seguros_de_Vidas':0,
                        'Aporte_1': 0,
                        'Aporte_Adicional':0,
                        'Otros_CCAF':0,
                        'Monto_Asignación_Familiar_Retroactiva':0,
                        'Monto_Reintegros_Asignación_Familiar':0,
                        'Monto_6_de_trabajador_no_afiliado_a_Isapre':0,
                        'Productos_financieros':DfHE.iloc[i]['MONTO'],
                        'afa':0,
                        })
        dtindex = pd.DataFrame(index)
        dtindex.to_csv("SalidaCVS/index.csv",encoding='utf-8')
    except Exception as e: logging.error(e)

# _____________________Task_________________________